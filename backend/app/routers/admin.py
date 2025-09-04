from typing import List
from fastapi import APIRouter, Depends, HTTPException, status, Response
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.auth import get_admin_user
from app.crud.crud_user import get_users, update_user
from app.crud.crud_question import update_question_status, get_questions_for_export
from app.schemas.schemas import User, UserUpdate, QuestionStatus
from app.models.models import User as UserModel
import openpyxl
from openpyxl.styles import Font, Alignment
import io

router = APIRouter(prefix="/admin", tags=["Administração"])

@router.get("/users", response_model=List[User])
def list_all_users(
    skip: int = 0,
    limit: int = 100,
    admin_user: UserModel = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """Listar todos os usuários"""
    return get_users(db, skip=skip, limit=limit)

@router.put("/users/{user_id}/role", response_model=User)
def update_user_role(
    user_id: int,
    user_update: UserUpdate,
    admin_user: UserModel = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """Atualizar role de usuário"""
    updated_user = update_user(db=db, user_id=user_id, user_update=user_update)
    if updated_user is None:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return updated_user

@router.put("/questions/{question_id}/status")
def update_question_status_admin(
    question_id: int,
    status: QuestionStatus,
    admin_user: UserModel = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """Atualizar status de questão (aprovar/rejeitar)"""
    updated_question = update_question_status(db=db, question_id=question_id, status=status)
    if updated_question is None:
        raise HTTPException(status_code=404, detail="Questão não encontrada")
    return {"message": f"Status da questão atualizado para {status}", "question_id": question_id}

@router.get("/export/excel")
def export_questions_excel(
    status: QuestionStatus = None,
    admin_user: UserModel = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """Exportar questões para Excel"""
    questions = get_questions_for_export(db, status=status)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questões de Geografia"
    
    # Cabeçalhos
    headers = [
        'ID Questão', 'Data', 'Usuário', 'Email Usuário', 'Tema Principal', 'Subtópico',
        'Enunciado', 'Tipo de questão', 'URL da imagem', 'Descrição da imagem',
        'Fonte da imagem', 'Nível Escolar', 'Alternativa A', 'Alternativa B',
        'Alternativa C', 'Alternativa D', 'Alternativa E', 'Resposta correta',
        'Texto Alternativa correta', 'Dica', 'Fonte/Referência bibliográfica', 'Status'
    ]
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar dados
    for row, question in enumerate(questions, 2):
        data = [
            question.id,
            question.data_cadastro.strftime('%d/%m/%Y'),
            question.user.username,
            question.user.email,
            question.tema_principal,
            question.subtopico,
            question.enunciado,
            question.tipo_questao,
            question.url_imagem or '',
            question.descricao_imagem or '',
            question.fonte_imagem or '',
            question.nivel_escolar,
            question.alternativa_a,
            question.alternativa_b,
            question.alternativa_c,
            question.alternativa_d,
            question.alternativa_e,
            question.resposta_correta,
            question.texto_alternativa_correta,
            question.dica or '',
            question.fonte_bibliografica or '',
            question.status
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em memória
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Retornar arquivo
    filename = f"questoes_geografia_{status or 'todas'}.xlsx"
    return Response(
        content=excel_buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

